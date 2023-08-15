import os
import cv2
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as xlImage
import FunctionsWithROIs as fp
from datetime import datetime
import tempfile
import BestImageFinder as bif


################# The only things ypu need to change ######################
# path to the images
mypath = r'G:\AI Engineering\Co-ops\Chitwan Singh\Image Characterization\Defect Testing\14100\Chatter'
 
# Final Folder Name
resultFolderName =  "Final Result"
################# The only things ypu need to change ######################


# current dateTime
now = datetime.now()

# convert to string
date_time_str = now.strftime("%Y-%m-%d %H_%M_%S")
print('DateTime String:', date_time_str)

# Filter out folders and select only image files
image_files = [
    f
    for f in os.listdir(mypath)
    if os.path.isfile(os.path.join(mypath, f))
    and f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp'))
]

# Sort the image files numerically
images = sorted(image_files, key=lambda x: int(''.join(filter(str.isdigit, x))))

# Create Excel Workbook
wb = openpyxl.Workbook()

# Select the active sheet
sheet1 = wb.active
sheet1.title = "No Edge Map" 

# Create a new sheet
sheet2 = wb.create_sheet(title="With Edge Map")

# Set header style
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
header_cell_styles = [header_font, header_alignment]

#Sheet 1 Header Font
sheet1['A1'].font = header_font
sheet1['B1'].font = header_font
sheet1['C1'].font = header_font
sheet1['D1'].font = header_font
sheet1['E1'].font = header_font
sheet1['F1'].font = header_font
sheet1['G1'].font = header_font
sheet1['H1'].font = header_font
sheet1['I1'].font = header_font

#Sheet 2 Header Font
sheet2['A1'].font = header_font
sheet2['B1'].font = header_font
sheet2['C1'].font = header_font
sheet2['D1'].font = header_font
sheet2['E1'].font = header_font
sheet2['F1'].font = header_font
sheet2['G1'].font = header_font
sheet2['H1'].font = header_font
sheet2['I1'].font = header_font
sheet2['J1'].font = header_font


# Apply header cell styles
for cell in sheet1[1]:
    for style in header_cell_styles:
        cell.alignment = style

# Writing Headers Sheet 1
sheet1['A1'] = 'IMAGE NAME'
sheet1['B1'] = 'SHARPNESS'
sheet1['C1'] = 'CONTRAST'
sheet1['D1'] = 'BRIGHTNESS'
sheet1['E1'] = 'SIGNAL-TO-NOISE (SNR)'
sheet1['F1'] = 'CONTRAST-TO-NOISE (CNR)'
sheet1['G1'] = 'DYNAMIC RANGE'
sheet1['H1'] = 'UNIFORMITY'
sheet1['I1'] = 'LINEARITY'

# Writing Headers Sheet 2
sheet2['A1'] = 'IMAGE NAME'
sheet2['B1'] = 'SHARPNESS'
sheet2['C1'] = 'CONTRAST'
sheet2['D1'] = 'BRIGHTNESS'
sheet2['E1'] = 'SIGNAL-TO-NOISE (SNR)'
sheet2['F1'] = 'CONTRAST-TO-NOISE (CNR)'
sheet2['G1'] = 'DYNAMIC RANGE'
sheet2['H1'] = 'UNIFORMITY'
sheet2['I1'] = 'LINEARITY'
sheet2['J1'] = 'EDGE MAP'

# Set alignment for header row
header_row_alignment = Alignment(horizontal='center', vertical='center')

# Sheet 1
sheet1['A1'].alignment = header_row_alignment
sheet1['B1'].alignment = header_row_alignment
sheet1['C1'].alignment = header_row_alignment
sheet1['D1'].alignment = header_row_alignment
sheet1['E1'].alignment = header_row_alignment
sheet1['F1'].alignment = header_row_alignment
sheet1['G1'].alignment = header_row_alignment
sheet1['H1'].alignment = header_row_alignment
sheet1['I1'].alignment = header_row_alignment

# Sheet 2
sheet2['A1'].alignment = header_row_alignment
sheet2['B1'].alignment = header_row_alignment
sheet2['C1'].alignment = header_row_alignment
sheet2['D1'].alignment = header_row_alignment
sheet2['E1'].alignment = header_row_alignment
sheet2['F1'].alignment = header_row_alignment
sheet2['G1'].alignment = header_row_alignment
sheet2['H1'].alignment = header_row_alignment
sheet2['I1'].alignment = header_row_alignment
sheet2['J1'].alignment = header_row_alignment

#Gte th eiage with the best Dynamic range
maxDynamicRange = 0 
maxDynamicRangeImage = ''
for i, imageName in enumerate(images, start=2):
    print(imageName)
    imagePath = os.path.join(mypath, imageName)

    # Find ROIs
    rois = fp.calculate_ROIs(imagePath)

    # Add image name to the excel sheet
    sheet1.cell(row=i, column=1, value=str(imageName))
    sheet2.cell(row=i, column=1, value=str(imageName))

    sharpness = fp.calculate_sharpness(imagePath)
    sheet1.cell(row=i, column=2, value=str(round(sharpness, 2)))
    sheet2.cell(row=i, column=2, value=str(round(sharpness, 2)))

    contrast = fp.calculate_contrast(imagePath, rois)
    if contrast >= 110:
        contrast = 0
    sheet1.cell(row=i, column=3, value=str(round(contrast, 2)))
    sheet2.cell(row=i, column=3, value=str(round(contrast, 2)))

    brightness = fp.calculate_brightness(imagePath, rois)
    sheet1.cell(row=i, column=4, value=str(round(brightness, 2)))
    sheet2.cell(row=i, column=4, value=str(round(brightness, 2)))

    snr = fp.calculate_snr(imagePath, rois)
    sheet1.cell(row=i, column=5, value=str(round(snr, 2)))
    sheet2.cell(row=i, column=5, value=str(round(snr, 2)))

    cnr = fp.calculate_cnr(imagePath, rois)
    sheet1.cell(row=i, column=6, value=str(round(cnr, 2)))
    sheet2.cell(row=i, column=6, value=str(round(cnr, 2)))

    dynamicRange = fp.calculate_dynamic_range(imagePath)
    if dynamicRange > maxDynamicRange:
        maxDynamicRange = dynamicRange
        maxDynamicRangeImage = imagePath
    sheet1.cell(row=i, column=7, value=str(round(dynamicRange, 2)))
    sheet2.cell(row=i, column=7, value=str(round(dynamicRange, 2)))

    uniformity = fp.calculate_uniformity(imagePath, rois)
    sheet1.cell(row=i, column=8, value=str(round(uniformity, 2)))
    sheet2.cell(row=i, column=8, value=str(round(uniformity, 2)))

    linearity = fp.calculate_linearity(imagePath, rois)
    sheet1.cell(row=i, column=9, value=str(round(linearity, 2)))
    sheet2.cell(row=i, column=9, value=str(round(linearity, 2)))

    image = fp.edge_detection(imagePath)

    # Save the image as a temporary file
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_img:
        temp_img_path = temp_img.name
        cv2.imwrite(temp_img_path, image)

        # Create an openpyxl image object from the temporary image file
        xl_image = xlImage(temp_img_path)

        # Insert the image into the specified column
        column_letter = openpyxl.utils.get_column_letter(10)
        sheet2.add_image(xl_image, f"{column_letter + str(i)}")

        # Close the temporary image file
        temp_img.close()


# Auto-adjust column widths in Sheet 1
for column in sheet1.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet1.column_dimensions[column_letter].width = adjusted_width

# Auto-adjust column widths in Sheet 2
for column in sheet2.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet2.column_dimensions[column_letter].width = adjusted_width

# Auto-adjust row heights only in Sheet 2(excluding the first row)
for row in sheet2.iter_rows(min_row=2):
    max_height = 0
    for cell in row:
        try:
            img_height = xl_image.height
            if img_height > max_height:
                max_height = img_height
        except:
            pass
    adjusted_height = max_height
    sheet2.row_dimensions[row[0].row].height = adjusted_height

result = os.makedirs(mypath + '\\' + resultFolderName + '\\')
# Save the workbook
wb.save(mypath + '\\' + resultFolderName + '\\' + date_time_str + ' Images Data.xlsx')


best_image_path = bif.find_best_image(mypath)

with open(mypath + '\\' + 'Best Images.txt', 'w') as f:
    f.write("Best image:" + best_image_path)
    f.write('\n')
    f.write("Best dynamic range image:" + maxDynamicRangeImage) 
print(" " * 20)
print("Best image:", best_image_path)
print("Best dynamic range image:", maxDynamicRangeImage)
print(" " * 20)
print("############## Task Completed ##############")
print(" " * 20)

